# -*- coding: utf-8 -*-
"""
WB Cashback:
- Бронза: Selenium -> проверка наличия строки прошлой недели, затем прямой HTTP запрос к API для XLSX.
- Сильвер: парсинг XLSX из бронзы и запись агрегатов.
"""

from __future__ import annotations

import base64
import json
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from zoneinfo import ZoneInfo

import requests
from dagster import Any, In, Nothing, Out, DynamicOut, DynamicOutput, get_dagster_logger, op
from openpyxl import load_workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from sqlalchemy import bindparam, text
from sqlalchemy.dialects.postgresql import BYTEA

# ─────────────────────────────────────────────────────────────────────────────
# Константы / селекторы
# ─────────────────────────────────────────────────────────────────────────────

CASHBACK_URL = (
    "https://seller.wildberries.ru/"
    "suppliers-mutual-settlements/reports-implementations/cashback-reports"
)

# Таблица в ЛК WB (классы с префиксом и хэшем)
TABLE_BODY_SELECTOR = 'div[class^="Custom-table__body"]'
ROW_SELECTOR = 'div[class^="Custom-table__row"]'
CELL_SELECTOR = 'div[class^="Custom-table__cell"]'

# API, которое фактически отдаёт отчёт
CASHBACK_API = (
    "https://seller-weekly-report.wildberries.ru"
    "/ns/realization-reports/suppliers-portal-analytics/api/v1/cashback/download"
)

# ─────────────────────────────────────────────────────────────────────────────
# Время
# ─────────────────────────────────────────────────────────────────────────────

def _msk_now_from_context(context) -> datetime:
    scheduled_iso = context.dagster_run.tags.get("dagster/scheduled_execution_time")
    raw_dt = datetime.fromisoformat(scheduled_iso) if scheduled_iso else datetime.now(timezone.utc)
    return raw_dt.astimezone(ZoneInfo("Europe/Moscow"))


def monday_of_previous_week(msk_now: datetime) -> datetime:
    monday_this_week = (msk_now - timedelta(days=msk_now.weekday())).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    return monday_this_week - timedelta(days=7)


# ─────────────────────────────────────────────────────────────────────────────
# 1) Находим строку прошлой недели и подготавливаем мету запуска/тайминги
# ─────────────────────────────────────────────────────────────────────────────

@op(
    out=DynamicOut(dict),
    required_resource_keys={"selenium_remote"},
    description=(
        "Находит строку прошлой недели, возвращает start/end + cookies + authorizev3, "
        "а также run_dttm/run_schedule_dttm/business_dttm."
    ),
)
def get_cashback_download_meta(context):
    log = get_dagster_logger()

    scheduled_iso = context.dagster_run.tags.get("dagster/scheduled_execution_time")
    run_schedule_dttm = (
        datetime.fromisoformat(scheduled_iso).astimezone(ZoneInfo("Europe/Moscow"))
        if scheduled_iso else None
    )

    run_dttm = datetime.now(timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
    business_dttm = run_dttm.replace(minute=0, second=0, microsecond=0) - timedelta(weeks=1)

    # Для выбора строки используем «плановое сейчас», если есть; иначе — фактическое
    ref_now = run_schedule_dttm or run_dttm
    target_monday = (ref_now - timedelta(days=ref_now.weekday() + 7)).date().isoformat()

    profiles = context.resources.selenium_remote.profiles
    log.debug(
        "[cashback] run_dttm=%s, run_schedule_dttm=%s, business_dttm=%s, target_monday=%s",
        run_dttm, run_schedule_dttm, business_dttm, target_monday
    )

    for token_id_str, profile_name in profiles.items():
        token_id = int(token_id_str)
        driver = context.resources.selenium_remote(token_id)
        try:
            log.debug(f"[{profile_name}] opening {CASHBACK_URL}")
            driver.get(CASHBACK_URL)

            cur = driver.current_url
            if "passport.wildberries.ru" in cur or "login" in cur.lower():
                raise RuntimeError(f"[{profile_name}] WB требует авторизацию (профиль не залогинен)")

            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, TABLE_BODY_SELECTOR))
            )
            WebDriverWait(driver, 60).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, ROW_SELECTOR))
            )

            rows = driver.find_elements(By.CSS_SELECTOR, ROW_SELECTOR)
            log.debug(f"[{profile_name}] rows found: {len(rows)}")
            if not rows:
                log.error(f"[cashback] {profile_name}: таблица пустая")
                continue

            # ищем точное совпадение по началу периода (1-й столбец)
            target_row = None
            start_txt = end_txt = None
            first_two = None

            for idx, row in enumerate(rows[:3]):
                cells = row.find_elements(By.CSS_SELECTOR, CELL_SELECTOR)
                vals = [c.text.strip() for c in cells[:2]]
                if idx == 0:
                    first_two = vals
                if len(vals) >= 2 and vals[0] == target_monday:
                    target_row, start_txt, end_txt = row, vals[0], vals[1]
                    break

            if not target_row:
                if first_two and len(first_two) >= 2:
                    target_row = rows[0]
                    start_txt, end_txt = first_two[0], first_two[1]
                    log.warning(
                        f"[cashback] {profile_name}: нет {target_monday}, берём верхнюю {start_txt}–{end_txt}"
                    )
                else:
                    log.error(f"[cashback] {profile_name}: не нашли валидные ячейки в первой строке")
                    continue

            selenium_cookies = {c["name"]: c["value"] for c in driver.get_cookies()}
            authv3 = driver.execute_script(
                "return window.localStorage.getItem('wb-eu-passport-v2.access-token');"
            )
            log.debug(
                f"[{profile_name}] cookies={list(selenium_cookies.keys())}; "
                f"authorizev3={'<set>' if authv3 else '<empty>'}"
            )

            yield DynamicOutput(
                {
                    "api_token_id": token_id,
                    "profile_name": profile_name,
                    "start": start_txt,
                    "end": end_txt,
                    # времена ранa:
                    "run_dttm": run_dttm.isoformat(),
                    "run_schedule_dttm": run_schedule_dttm.isoformat() if run_schedule_dttm else None,
                    "business_dttm": business_dttm.isoformat(),
                    # для скачивания:
                    "cookies": selenium_cookies,
                    "authv3": authv3,
                },
                mapping_key=str(token_id),
            )

        finally:
            try:
                driver.quit()
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────────────────────
# 2) HTTP GET XLSX + запись в бронзу (BYTEA) с новыми полями времени
# ─────────────────────────────────────────────────────────────────────────────

@op(
    ins={"cashback_meta": In(dict)},
    out=Out(Any),
    required_resource_keys={"postgres"},
    description=(
        "Качает XLSX по API cashback/download (dateFrom/dateTo) и пишет в bronze.wb_www_cashback_reports_1w. "
        "Заполняет run_dttm/run_schedule_dttm/business_dttm/request_dttm/receive_dttm/response_dttm."
    ),
)
async def download_and_write_cashback_xlsx(context, cashback_meta: dict):
    log = get_dagster_logger()
    session_maker = context.resources.postgres

    api_token_id   = int(cashback_meta["api_token_id"])
    profile_name   = cashback_meta["profile_name"]
    start_txt      = cashback_meta["start"]
    end_txt        = cashback_meta["end"]

    # времена ранa пришли из первого опса — используем как есть, чтобы были едиными
    run_dttm            = datetime.fromisoformat(cashback_meta["run_dttm"])
    run_schedule_dttm_s = cashback_meta.get("run_schedule_dttm")
    run_schedule_dttm   = datetime.fromisoformat(run_schedule_dttm_s) if run_schedule_dttm_s else None
    business_dttm       = datetime.fromisoformat(cashback_meta["business_dttm"])

    # cookies / headers
    s = requests.Session()
    for name, value in cashback_meta["cookies"].items():
        s.cookies.set(name, value)

    supplier_id = (
        cashback_meta["cookies"].get("x-supplier-id")
        or cashback_meta["cookies"].get("x-supplier-id-external")
        or ""
    )
    params = {"dateFrom": start_txt, "dateTo": end_txt}
    url = CASHBACK_API
    headers = {
        "accept": "*/*",
        "authorizev3": cashback_meta.get("authv3") or "",
        "X-Supplier-Id": supplier_id,
        "origin": "https://seller.wildberries.ru",
        "referer": "https://seller.wildberries.ru/",
        "user-agent": "Mozilla/5.0",
        "wb-seller-lk": "true",
    }

    # request_dttm — прямо перед запросом
    request_dttm = datetime.now(timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
    log.debug(
        f"[{profile_name}] GET {url} params={params} "
        f"(run_dttm={run_dttm}, run_schedule_dttm={run_schedule_dttm}, business_dttm={business_dttm}, "
        f"request_dttm={request_dttm})"
    )

    resp = s.get(url, params=params, headers=headers, timeout=120, stream=True)
    # receive_dttm — сразу после ответа
    receive_dttm = datetime.now(timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
    log.debug(
        f"[{profile_name}] status={resp.status_code} ctype={resp.headers.get('content-type')} "
        f"clen={resp.headers.get('content-length')} receive_dttm={receive_dttm}"
    )
    resp.raise_for_status()

    # response_dttm = считаем равным receive_dttm (в WB часто нет точного server-time)
    response_dttm = receive_dttm

    content_type_hdr = resp.headers.get("content-type", "")
    file_name = "cashback.xlsx"
    content: bytes

    if "application/json" in content_type_hdr.lower():
        data = resp.json()
        node = data.get("data", data)
        b64 = node.get("file") or node.get("content") or ""
        fn  = node.get("fileName") or node.get("filename") or ""
        if not b64:
            raise RuntimeError(f"API вернул JSON без файла: keys={list(node.keys())}")
        content = base64.b64decode(b64)
        if fn:
            file_name = fn
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        http_code = 200
    else:
        content = resp.content
        http_code = resp.status_code
        cd = resp.headers.get("content-disposition", "")
        m  = re.search(r'filename="?([^"]+)"?', cd or "", re.I)
        if m:
            file_name = m.group(1)
        content_type = resp.headers.get("content-type") or \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if content[:2] != b"PK":
        log.warning(f"[{profile_name}] получен не ZIP/XLSX (magic={content[:4]!r}), сохраняем как есть")

    run_uuid = context.run_id
    request_uuid = str(uuid.uuid4())

    request_parameters = {
        "start": start_txt,
        "end": end_txt,
        "profile": profile_name,
        "url": url,
        "params": params,
        "file_name": file_name,
        "content_type": content_type,
    }

    log.debug(
        f"[{profile_name}] insert bronze: api_token_id={api_token_id}, bytes={len(content)}, "
        f"file={file_name}, response_code={http_code}"
    )

    stmt = text("""
        INSERT INTO bronze.wb_www_cashback_reports_1w (
            api_token_id, run_uuid,
            run_dttm, run_schedule_dttm, business_dttm,
            request_uuid, request_dttm, receive_dttm,
            request_parameters, request_body,
            response_dttm, response_code, response_body
        )
        VALUES (
            :api_token_id, :run_uuid,
            :run_dttm, :run_schedule_dttm, :business_dttm,
            :request_uuid, :request_dttm, :receive_dttm,
            :request_parameters, :request_body,
            :response_dttm, :response_code, :response_body
        )
    """).bindparams(bindparam("response_body", type_=BYTEA))

    async with session_maker() as sdb:
        await sdb.execute(
            stmt,
            {
                "api_token_id": api_token_id,
                "run_uuid": run_uuid,
                # новые поля времени:
                "run_dttm": run_dttm,
                "run_schedule_dttm": run_schedule_dttm,
                "business_dttm": business_dttm,
                # request/response тайминги:
                "request_uuid": request_uuid,
                "request_dttm": request_dttm,
                "receive_dttm": receive_dttm,
                "request_parameters": json.dumps(request_parameters, ensure_ascii=False),
                "request_body": None,
                "response_dttm": response_dttm,
                "response_code": http_code,
                "response_body": content,
            },
        )
        await sdb.commit()

    return {
        "api_token_id": api_token_id,
        "start": start_txt,
        "end": end_txt,
        "bytes": len(content),
        "file_name": file_name,
        "run_dttm": run_dttm.isoformat(),
        "run_schedule_dttm": run_schedule_dttm.isoformat() if run_schedule_dttm else None,
        "business_dttm": business_dttm.isoformat(),
        "request_dttm": request_dttm.isoformat(),
        "receive_dttm": receive_dttm.isoformat(),
    }



# ─────────────────────────────────────────────────────────────────────────────
# Helpers для парсинга XLSX
# ─────────────────────────────────────────────────────────────────────────────

def _to_decimal(x) -> Decimal | None:
    if x is None:
        return None
    if isinstance(x, (int, float, Decimal)):
        try:
            return Decimal(str(x))
        except InvalidOperation:
            return None
    s = str(x).strip().replace("\u00A0", "").replace(" ", "")
    if s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_date(x) -> date | None:
    if x is None:
        return None
    if isinstance(x, date) and not isinstance(x, datetime):
        return x
    if isinstance(x, datetime):
        return x.date()
    s = str(x).strip().replace("\u00A0", " ")
    try:
        if "." in s:  # DD.MM.YYYY
            d, m, y = s.split(".")
            return date(int(y), int(m), int(d))
        return date.fromisoformat(s.split("T")[0])  # YYYY-MM-DD
    except Exception:
        return None


def _norm_key(s: str) -> str:
    """Нормализуем русские подписи из файла до ключа словаря."""
    if s is None:
        return ""
    t = s.lower().replace("ё", "е").replace("\u00A0", " ")
    return " ".join(t.split())


# Прямые соответствия «подпись → имя поля silver»
LABEL_MAP = {
    "дата начала периода": "period_start_date",
    "дата окончания периода": "period_end_date",
    "сумма начислений кэшбека за период": "cashback_provided",
    "сумма начислений кешбека за период": "cashback_provided",
    "сумма погашений кэшбека за период": "cashback_used",
    "сумма погашений кешбека за период": "cashback_used",
    "сумма сгоревшего кэшбека за период": "cashback_expired",
    "сумма сгоревшего кешбека за период": "cashback_expired",
    "сумма возвращенного кэшбека за период": "cashback_returned",
    "сумма возвращенного кешбека за период": "cashback_returned",
    "баланс/сальдо на окончание периода": "cashback_final_balance",
}


def _parse_cashback_file(xlsx_bytes: bytes) -> dict:
    """
    Жёсткий парсер под файл «Отчёт по кэшбэкам …».
    В первом листе строки вида: <Подпись> | <Значение>.
    """
    log = get_dagster_logger()
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    result = {
        "period_start_date": None,
        "period_end_date": None,
        "cashback_provided": None,
        "cashback_used": None,
        "cashback_expired": None,
        "cashback_returned": None,
        "cashback_final_balance": None,
    }

    matched = 0
    for row in ws.iter_rows(values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        try:
            label_idx = next(i for i, v in enumerate(row) if v not in (None, ""))
        except StopIteration:
            continue

        label = _norm_key(str(row[label_idx]))
        field = LABEL_MAP.get(label)
        if not field:
            continue

        # значение — первая непустая ячейка правее подписи
        val = next((v for v in row[label_idx + 1 :] if v not in (None, "")), None)

        parsed = _to_date(val) if field.startswith("period_") else _to_decimal(val)
        result[field] = parsed
        matched += 1
        log.debug(f"[xlsx] matched '{label}' -> {field} = {parsed}")

    log.debug(f"[xlsx] parsed fields: {matched}/{len(result)}")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 3) Silver: парсим XLSX из бронзы и пишем агрегаты
# ─────────────────────────────────────────────────────────────────────────────

@op(
    ins={"start": In(Nothing)},
    out=Out(Any),
    required_resource_keys={"postgres"},
    description="Парсит XLSX из bronze.wb_www_cashback_reports_1w и пишет агрегаты в silver.wb_www_cashback_reports_1w",
)
async def upsert_cashback_silver_from_bronze(context):
    log = get_dagster_logger()
    session_maker = context.resources.postgres
    run_uuid = context.run_id

    # вытаскиваем ТОЛЬКО записи текущего run, и тащим business_dttm
    select_sql = text("""
        SELECT
            b.api_token_id        AS company_id,
            b.request_uuid,
            b.inserted_at,
            b.response_dttm,
            b.business_dttm,
            b.request_parameters,
            b.response_body
        FROM bronze.wb_www_cashback_reports_1w b
        LEFT JOIN silver.wb_www_cashback_reports_1w s
               ON s.company_id = b.api_token_id
              AND s.business_dttm = b.business_dttm
        WHERE b.run_uuid = :run_uuid
          AND s.company_id IS NULL
        ORDER BY b.inserted_at
    """)

    async with session_maker() as sess:
        rows = (await sess.execute(select_sql, {"run_uuid": run_uuid})).fetchall()

    log.info(f"[silver] bronze rows to process (current run): {len(rows)}")
    if not rows:
        return {"inserted": 0}

    inserted = 0
    async with session_maker() as sess:
        for (company_id, request_uuid, inserted_at, response_dttm, business_dttm,
             req_params, blob) in rows:
            try:
                if not blob or blob[:2] != b"PK":
                    raise ValueError("response_body не похож на XLSX (нет сигнатуры PK)")

                metrics = _parse_cashback_file(blob)

                # подстраховка дат из request_parameters (если не нашлись в файле)
                params = req_params or {}
                if isinstance(params, str):
                    params = json.loads(params)
                if not metrics["period_start_date"]:
                    metrics["period_start_date"] = _to_date(params.get("start") or params.get("period_start_date"))
                if not metrics["period_end_date"]:
                    metrics["period_end_date"]   = _to_date(params.get("end")   or params.get("period_end_date"))

                upsert_sql = text("""
                    INSERT INTO silver.wb_www_cashback_reports_1w (
                        company_id, business_dttm, request_uuid,
                        inserted_at, response_dttm,
                        period_start_date, period_end_date,
                        cashback_provided, cashback_used, cashback_expired,
                        cashback_returned, cashback_final_balance
                    )
                    VALUES (
                        :company_id, :business_dttm, :request_uuid,
                        :inserted_at, :response_dttm,
                        :period_start_date, :period_end_date,
                        :cashback_provided, :cashback_used, :cashback_expired,
                        :cashback_returned, :cashback_final_balance
                    )
                    ON CONFLICT (company_id, business_dttm) DO UPDATE SET
                        request_uuid         = EXCLUDED.request_uuid,
                        response_dttm        = EXCLUDED.response_dttm,
                        period_start_date    = EXCLUDED.period_start_date,
                        period_end_date      = EXCLUDED.period_end_date,
                        cashback_provided    = EXCLUDED.cashback_provided,
                        cashback_used        = EXCLUDED.cashback_used,
                        cashback_expired     = EXCLUDED.cashback_expired,
                        cashback_returned    = EXCLUDED.cashback_returned,
                        cashback_final_balance = EXCLUDED.cashback_final_balance
                """)

                await sess.execute(
                    upsert_sql,
                    {
                        "company_id": company_id,
                        "business_dttm": business_dttm,
                        "request_uuid": request_uuid,
                        "inserted_at": inserted_at,
                        "response_dttm": response_dttm,
                        "period_start_date": metrics["period_start_date"],
                        "period_end_date": metrics["period_end_date"],
                        "cashback_provided": metrics["cashback_provided"],
                        "cashback_used": metrics["cashback_used"],
                        "cashback_expired": metrics["cashback_expired"],
                        "cashback_returned": metrics["cashback_returned"],
                        "cashback_final_balance": metrics["cashback_final_balance"],
                    },
                )
                inserted += 1
            except Exception as e:
                log.error(f"[silver] пропускаем {company_id}/{business_dttm}: {e}")

        await sess.commit()

    log.info(f"[silver] inserted/updated rows: {inserted}")
    return {"inserted": inserted}
